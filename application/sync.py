import serial.tools.list_ports
from comm import Comm
from plots import update_pid_plot
from diagnostics import log, diagnostics_path
import time
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import os

class Sync:
    def __init__(self, parent, frequency):

        self.parent = parent
        self.frequency = frequency
        
        self.comm = None

        self.port_list = []
        self.ports()

        self.monitor()

        self.health_status = False
        self.health()

        self.performance()
        
        self.pid_calibration_status = False
        self.pid_calibration_status_old = False
        self.pid_calibration_start_time = None
    
        self.pid_experiment_status = False
        self.pid_experiment_status_old  = False
        self.pid_experiment_start_time = None

        self.pid_experiment_duration_ms = 0
        self.pid_target_load_lbf = 0

        self.pid_plot_time = []
        self.pid_plot_load = []

        self.control()

        self.data_frequency_ms = 0
        self.experiment_data_start_time = 0
        self.experiment_data_count = 0
        self.save_experiment_data()

    def ports(self):
        raw_ports = [p.device for p in serial.tools.list_ports.comports()]
        new_ports = raw_ports if raw_ports else [""]

        if self.comm and self.comm.port not in new_ports:
            log("#501", "Application", "Port has been disappeared.")
            try:
                self.comm.disconnect()
            except Exception:
                pass

            self.comm = None
            self.parent.settings_frame.r2_c1.configure(text="Connect")
            self.parent.settings_frame.r2_c1.unbind("<Enter>")
            self.parent.settings_frame.r2_c1.unbind("<Leave>")
            
        if new_ports != self.port_list:
            self.port_list = new_ports

            self.parent.settings_frame.r1_c1.configure(values=self.port_list) 

            if self.parent.settings_frame.r1_c1.get() not in self.port_list:
                self.parent.settings_frame.r1_c1.set(self.port_list[0])

            if self.comm:
                current_port = self.comm.port
                if current_port not in new_ports:
                    log("#502", "Application", "Port is not accessible.")
                    self.comm.disconnect()
                    self.comm = None
                    self.parent.settings_frame.r2_c1.configure(text="Connect")
                    self.parent.settings_frame.r2_c1.unbind("<Enter>")
                    self.parent.settings_frame.r2_c1.unbind("<Leave>")

        self.parent.after(self.frequency, self.ports)

    def connection(self):
        if self.comm:
            try:
                self.comm.disconnect()
                self.comm = None
                self.parent.settings_frame.r2_c1.configure(text="Connect")
                self.parent.settings_frame.r2_c1.unbind("<Enter>")
                self.parent.settings_frame.r2_c1.unbind("<Leave>")
            except Exception:
                self.comm = None
                log("#503", "Application", "Serial cannot be terminated.")
        else:
            port = self.parent.settings_frame.r1_c1.get()
            if port:
                try:
                    baudrate = int(self.parent.config.get("comm_baudrate", "115200"))
                    timeout = float(self.parent.config.get("comm_timeout", "1"))
                    frequency = int(self.parent.config.get("comm_frequency", "100"))

                    self.comm = Comm(port=port, baudrate=baudrate, timeout=timeout, frequency=frequency)
                    
                    self.comm.send_config(self.parent.config)
                    
                    self.parent.settings_frame.r2_c1.configure(text="Connected")

                    def on_enter(event):
                        self.parent.settings_frame.r2_c1.configure(text="Disconnect")
                    def on_leave(event):
                        self.parent.settings_frame.r2_c1.configure(text="Connected")

                    self.parent.settings_frame.r2_c1.bind("<Enter>", on_enter)
                    self.parent.settings_frame.r2_c1.bind("<Leave>", on_leave)

                except Exception:
                    self.comm = None

    def monitor(self):
        if self.comm:
            try:
                hdr, bfr = self.comm.serial_monitor()
                if hdr == b'\xCC':
                    self.feedback(bfr)
                elif hdr == b'\xDD':
                    self.data(bfr)
            except Exception:
                pass
        self.parent.after(self.frequency, self.monitor)

    def feedback(self, bfr):
        if bfr:
            if bfr == b'\x10\x00':
                log("#504", "Controller", "Motors have been halted.")
                self.pid_calibration_status = False
                self.pid_experiment_status = False
            elif bfr == b'\x11\x00':
                log("#505", "Controller", "Rotary Motor has been rotated by 90-deg.")
            elif bfr == b'\x12\x00':
                log("#506", "Controller", "Rotary Motor has been sent to home.")
            elif bfr == b'\x13\x00':
                log("#507", "Controller", "Z-Axis Motors have been sent to home.")
            elif bfr == b'\x14\x00':
                log("#508", "Controller", "Motors have been moved by given distance.")
            elif bfr == b'\x14\x01':
                log("#509", "Controller", "Wrong size of degree unit has been received to move Rotary Motor by.")
            elif bfr == b'\x14\x02':
                log("#510", "Controller", "Wrong size of step unit has been received to move Right Z-Axis Motor by.")
            elif bfr == b'\x14\x03':
                log("#511", "Controller", "Wrong size of step unit has been received to move Left Z-Axis Motor by.")
            elif bfr == b'\x14\x04':
                log("#512", "Controller", "Wrong size of in unit has been received to move Right Z-Axis Motor by.")
            elif bfr == b'\x14\x05':
                log("#513", "Controller", "Wrong size of in unit has been received to move Left Z-Axis Motor by.")
            elif bfr == b'\x14\x06':
                log("#514", "Controller", "Wrong unit type has been received to move Z-Axis Motors by.")
            elif bfr == b'\x15\x00':
                log("#515", "Controller", "Motors have been moved to given position.")
            elif bfr == b'\x15\x01':
                log("#516", "Controller", "Wrong size of degree unit has been received to move Rotary Motor to.")
            elif bfr == b'\x15\x02':
                log("#517", "Controller", "Wrong size of step unit has been received to move Right Z-Axis Motor to.")
            elif bfr == b'\x15\x03':
                log("#518", "Controller", "Wrong size of step unit has been received to move Left Z-Axis Motor to.")
            elif bfr == b'\x15\x04':
                log("#519", "Controller", "Wrong size of in unit has been received to move Right Z-Axis Motor to.")
            elif bfr == b'\x15\x05':
                log("#520", "Controller", "Wrong size of in unit has been received to move Left Z-Axis Motor to.")
            elif bfr == b'\x15\x06':
                log("#521", "Controller", "Wrong unit type has been received to move Z-Axis Motors to.")
            elif bfr == b'\x16\x00':
                log("#522", "Controller", "Rotary Motor position has been set to home.")
            elif bfr == b'\x17\x00':
                log("#523", "Controller", "Right Z-Axis Motor position has been set to home.")
            elif bfr == b'\x18\x00':
                log("#524", "Controller", "Left Z-Axis Motor position has been set to home.")
            elif bfr == b'\x20\x00':
                log("#525", "Controller", "Experiment has been initiated.")
            elif bfr == b'\x20\x01':
                log("#526", "Controller", "Wrong size of Rotary Motor Speed has been received for experiment.")
            elif bfr == b'\x20\x02':
                log("#527", "Controller", "Wrong size of Target Load has been received for experiment.")
            elif bfr == b'\x20\x03':
                log("#528", "Controller", "Wrong size of Experiment Duration has been received for experiment.")
            elif bfr == b'\x20\x04':
                log("#529", "Controller", "The device calibration has been initiated for target load.")
                self.pid_calibration_status = True
            elif bfr == b'\x20\x05':
                log("#530", "Controller", "The device has been calibrated successfully for target load.")
                self.pid_calibration_status = False
            elif bfr == b'\x20\x06':
                log("#531", "Controller", "Experiment has been started.")
                self.pid_experiment_status = True
            elif bfr == b'\x20\x07':
                log("#532", "Controller", "Experiment has been completed successfully.")
                self.pid_experiment_status = False
            elif bfr == b'\x21\x00':
                log("#533", "Controller", "Experiment has been terminated.")
                self.pid_calibration_status = False
                self.pid_experiment_status = False
            elif bfr == b'\x99\x00':
                log("#534", "Controller", "Motors have been halted due to overload protection.")
                self.pid_calibration_status = False
                self.pid_experiment_status = False
            else:
                pass
                
    def data(self, bfr):
        if bfr:
            for key, value in bfr.items():
                if key in self.parent.data:
                    self.parent.data[key]["value"] = value

    def health(self):
        if self.comm:
            self.parent.health_frame.r0_c1.configure(text="Connected ⬤", text_color="green")

            if self.parent.data['right_loadcell_load_lbf']['value'] <= float(self.parent.config.get("right_loadcell_max_load_lbf")):
                self.parent.health_frame.r4_c1.configure(text="Active ⬤", text_color="green")
            else:
                self.parent.health_frame.r4_c1.configure(text="Overload ⬤", text_color="red")
            
            if self.parent.data['left_loadcell_load_lbf']['value'] <= float(self.parent.config.get("left_loadcell_max_load_lbf")):
                self.parent.health_frame.r5_c1.configure(text="Active ⬤", text_color="green")
            else:
                self.parent.health_frame.r5_c1.configure(text="Overload ⬤", text_color="red")

            if self.parent.data['rotary_motor_speed_step_s']['value'] != 0:
                self.parent.health_frame.r1_c1.configure(text="Busy ⬤", text_color="red")
            else:
                self.parent.health_frame.r1_c1.configure(text="Idle ⬤", text_color="green")

            if self.parent.data['right_z_axis_motor_speed_step_s']['value'] != 0:
                self.parent.health_frame.r2_c1.configure(text="Busy ⬤", text_color="red")
            else:
                self.parent.health_frame.r2_c1.configure(text="Idle ⬤", text_color="green")

            if self.parent.data['left_z_axis_motor_speed_step_s']['value'] != 0:
                self.parent.health_frame.r3_c1.configure(text="Busy ⬤", text_color="red")
            else:
                self.parent.health_frame.r3_c1.configure(text="Idle ⬤", text_color="green")
            
        else:

            self.parent.health_frame.r0_c1.configure(text="Not Connected ⬤", text_color="gray14")
            self.parent.health_frame.r1_c1.configure(text="Inactive ⬤", text_color="gray14")
            self.parent.health_frame.r2_c1.configure(text="Inactive ⬤", text_color="gray14")
            self.parent.health_frame.r3_c1.configure(text="Inactive ⬤", text_color="gray14")
            self.parent.health_frame.r4_c1.configure(text="Inactive ⬤", text_color="gray14")
            self.parent.health_frame.r5_c1.configure(text="Inactive ⬤", text_color="gray14")

        self.parent.after(self.frequency, self.health)

    def performance(self):
        if self.comm:
            self.parent.performance_frame.r1_c1.configure(text=f"{self.parent.data['rotary_motor_speed_rpm']['value']:.2f}")
            self.parent.performance_frame.r2_c1.configure(text=f"{self.parent.data['rotary_motor_position_deg']['value']:.2f}")
            self.parent.performance_frame.r3_c1.configure(text=f"{self.parent.data['passive_roller_position_in']['value']:.2f}")
            self.parent.performance_frame.r4_c1.configure(text=f"{self.parent.data['right_loadcell_load_lbf']['value']:.2f}")
            self.parent.performance_frame.r5_c1.configure(text=f"{self.parent.data['left_loadcell_load_lbf']['value']:.2f}")
            self.parent.performance_frame.r6_c1.configure(text=f"{self.parent.data['total_load_lbf']['value']:.2f}")
            self.parent.performance_frame.r7_c1.configure(text=f"{self.parent.data['thermocouple_flow_temperature_c']['value']:.2f}")

            if self.parent.advanced_performance is not None and self.parent.advanced_performance.winfo_exists():
                for key, entry in self.parent.data.items():
                    excluded_keys = {"pid_tuning_mode", "pid_overshoot_lbf","pid_overshoot_percent", "pid_settling_time_s", "pid_load_error_running_avg_percent"}
                    if key in excluded_keys:
                        continue
                    else: 
                        value = getattr(self.parent.advanced_performance, f"{key}_value", None)
                        value.configure(text=f"{entry['value']:.2f}")
        else:
            pass
            
        self.parent.after(self.frequency, self.performance)

    def control(self):
        if self.comm:
            
            if self.parent.PID_results is not None and self.parent.PID_results.winfo_exists():
                self.parent.PID_results.r0_c1.configure(text=f"{self.parent.data['pid_overshoot_lbf']['value']:.2f}")
                self.parent.PID_results.r1_c1.configure(text=f"{self.parent.data['pid_overshoot_percent']['value']:.2f}")
                self.parent.PID_results.r2_c1.configure(text=f"{self.parent.data['pid_settling_time_s']['value']:.2f}")
                self.parent.PID_results.r3_c1.configure(text=f"{self.parent.data['pid_load_error_running_avg_percent']['value']:.2f}")

            if self.pid_calibration_status and not self.pid_calibration_status_old:
                self.pid_calibration_start_time = time.time()
                self.pid_plot_time.clear()
                self.pid_plot_load.clear()

                if self.pid_target_load_lbf != 0:
                    self.parent.pid_frame.open_PID_results()

            if self.pid_calibration_status:
                calibration_elapsed_time = time.time() - self.pid_calibration_start_time
                self.pid_plot_time.append(calibration_elapsed_time)
                self.pid_plot_load.append(self.parent.data["total_load_lbf"]["value"])
                if self.pid_target_load_lbf != 0:
                    update_pid_plot(self.parent.pid_frame.ax, (self.pid_plot_time, self.pid_plot_load), target=float(self.pid_target_load_lbf), pid_tuning_mode="Conservative" if self.parent.data["pid_tuning_mode"]["value"] == 0 else "Aggressive")
                self.parent.pid_frame.canvas_pid.draw()
                
            self.pid_calibration_status_old = self.pid_calibration_status

            if self.pid_experiment_status and not self.pid_experiment_status_old:
                self.pid_experiment_start_time = time.time()
                self.pid_plot_time.clear()
                self.pid_plot_load.clear()

            if self.pid_experiment_status:
                experiment_elapsed_time = time.time() - self.pid_experiment_start_time
                hh, rem = divmod(int(experiment_elapsed_time), 3600)
                mm, ss = divmod(rem, 60)
                self.parent.control_frame.r12_c1.configure(text=f"{hh:02d}:{mm:02d}:{ss:02d}")
                self.parent.control_frame.r13_c1.configure(text=self.parent.data["total_revolution_rev"]["value"])
                self.pid_plot_time.append(experiment_elapsed_time)
                self.pid_plot_load.append(self.parent.data["total_load_lbf"]["value"])
                if self.pid_target_load_lbf != 0:
                    update_pid_plot(self.parent.pid_frame.ax, (self.pid_plot_time, self.pid_plot_load), target=float(self.pid_target_load_lbf), pid_tuning_mode="Conservative" if self.parent.data["pid_tuning_mode"]["value"] == 0 else "Aggressive")
                self.parent.pid_frame.canvas_pid.draw()
            else:
                self.parent.control_frame.r12_c1.configure(text="00:00:00")

            if not self.pid_experiment_status and self.pid_experiment_status_old:
                self.parent.control_frame.r11_c1.set("Stop")
                self.experiment_data_count = 0

            self.pid_experiment_status_old = self.pid_experiment_status

        else:   
            pass
            
        self.parent.after(self.frequency, self.control)

    def save_experiment_data(self):
        if (self.comm and self.pid_experiment_status and self.parent.control_frame.experiment_directory):
            experiment_name = self.parent.control_frame.r9_c1.get()
            if not experiment_name.endswith(".xlsx"):
                experiment_name += ".xlsx"
            
            experiment_path = os.path.join(self.parent.control_frame.experiment_directory, experiment_name)
            
            if experiment_path:
                try:
                    self.data_frequency_ms = (int)(self.parent.control_frame.r7_c1.get())

                    if self.experiment_data_count == 0:
                        wb = Workbook()

                        # Sheet 1: Data
                        data_sheet = wb.active
                        data_sheet.title = "Data"
                        
                        data_sheet_selected_vars = [entry["label"] for key, entry in self.parent.data.items() if entry.get("state", 0) == 1]
                        data_sheet_header = ["Date", "Time", "Duration (s)"] + data_sheet_selected_vars
                        data_sheet.append(data_sheet_header)
                        
                        # Sheet 2: Configuration
                        config_sheet = wb.create_sheet(title="Configuration")
                        config_sheet_variables = [
                            ["pid_rotary_motor_direction", self.parent.control_frame.r1_c1.get()],
                            ["pid_rotary_motor_speed_rpm", self.parent.control_frame.r2_c1.get()],
                            ["pid_target_load_lbf", self.parent.control_frame.r4_c1.get()],
                            ["pid_experiment_duration_hh_mm_ss", self.parent.control_frame.r5_c1.get()],
                            ["data_frequency_ms", self.parent.control_frame.r7_c1.get()],
                            ["pid_rotary_motor_ramplen_step", self.parent.config["pid_rotary_motor_ramplen_step"]],
                            ["pid_z_axis_motor_max_speed_rpm", self.parent.config["pid_z_axis_motor_max_speed_rpm"]],
                            ["pid_z_axis_motor_ramplen_step", self.parent.config["pid_z_axis_motor_ramplen_step"]],
                            ["pid_calibration_stable_required_time_ms", self.parent.config["pid_calibration_stable_required_time_ms"]],
                            ["pid_calibration_load_error_percent", self.parent.config["pid_calibration_load_error_percent"]],
                            ["pid_tuning_mode_limit_percent", self.parent.config["pid_tuning_mode_limit_percent"]],
                            ["pid_conservative_k_p", self.parent.config["pid_conservative_k_p"]],
                            ["pid_conservative_k_i", self.parent.config["pid_conservative_k_i"]],
                            ["pid_conservative_k_d",self.parent.config["pid_conservative_k_d"]],
                            ["pid_aggressive_k_p", self.parent.config["pid_aggressive_k_p"]],
                            ["pid_aggressive_k_i", self.parent.config["pid_aggressive_k_i"]],
                            ["pid_aggressive_k_d", self.parent.config["pid_aggressive_k_d"]],
                            ["go_home_z_axis_motors_bounce_step", self.parent.config["go_home_z_axis_motors_bounce_step"]],
                            ["right_loadcell_zero_offset_lbf", self.parent.config["right_loadcell_zero_offset_lbf"]],
                            ["right_loadcell_scale_factor", self.parent.config["right_loadcell_scale_factor"]],
                            ["right_loadcell_max_load_lbf", self.parent.config["right_loadcell_max_load_lbf"]],
                            ["left_loadcell_zero_offset_lbf", self.parent.config["left_loadcell_zero_offset_lbf"]],
                            ["left_loadcell_scale_factor", self.parent.config["left_loadcell_scale_factor"]],
                            ["left_loadcell_max_load_lbf", self.parent.config["left_loadcell_max_load_lbf"]],
                        ]

                        for variables in config_sheet_variables:
                            config_sheet.append(variables)
                    
                        wb.save(experiment_path)
                        wb.close()
                    
                        now = datetime.now()
                        now_ms = now.microsecond

                        if now_ms < 500:
                            rounded_time = now.replace(microsecond=0)
                        else:
                            rounded_time = now.replace(microsecond=0) + timedelta(seconds=1)

                        date_now = now.strftime("%Y-%m-%d")
                        time_now = rounded_time.strftime("%H:%M:%S.%f")[:-5]
                        self.experiment_data_start_time = rounded_time
                        duration_s = 0
                    
                    else:
                        # Subsequent times - use frequency-based calculation
                        duration_ms = self.experiment_data_count * self.data_frequency_ms
                        current_time = self.experiment_data_start_time + timedelta(milliseconds=duration_ms)
                        
                        date_now = current_time.strftime("%Y-%m-%d")
                        time_now = current_time.strftime("%H:%M:%S.%f")[:-5]
                        duration_s = duration_ms / 1000

                    # Sheet 1: Data
                    wb = load_workbook(experiment_path)
                    data_sheet = wb["Data"]

                    data_sheet_selected_values = [entry["value"] for key, entry in self.parent.data.items() if entry.get("state", 0) == 1]
                    data_row = [date_now, time_now, duration_s] + data_sheet_selected_values
                    data_sheet.append(data_row)

                    self.experiment_data_count += 1

                    wb.save(experiment_path)
                    wb.close()

                except Exception:
                    pass
            
        self.parent.after(self.data_frequency_ms if self.data_frequency_ms > 0 else self.frequency, self.save_experiment_data)

    def halt_motors(self):
        if self.comm:
            self.comm.halt_motors()
        else:
            pass

    def rotate_90deg_rotary_motor(self):
        if self.comm:
            self.comm.rotate_90deg_rotary_motor()
        else:
            pass

    def go_home_rotary_motor(self):
        if self.comm:
            self.comm.go_home_rotary_motor()
        else:
            pass

    def go_home_z_axis_motors(self):
        if self.comm:
            self.comm.go_home_z_axis_motors()
        else:
            pass
    
    def move_motors_by(self):
        if self.comm:
            unit = self.parent.advanced_control.r5_c1.get()
            rotary_motor = self.parent.advanced_control.r6_c1.get()
            right_z_axis_motor = self.parent.advanced_control.r7_c1.get()
            left_z_axis_motor = self.parent.advanced_control.r8_c1.get()
            self.comm.move_motors_by(unit, rotary_motor, right_z_axis_motor, left_z_axis_motor)
        else:
            pass

    def move_motors_to(self):
        if self.comm:
            unit = self.parent.advanced_control.r10_c1.get()
            rotary_motor = self.parent.advanced_control.r11_c1.get()
            right_z_axis_motor = self.parent.advanced_control.r12_c1.get()
            left_z_axis_motor = self.parent.advanced_control.r13_c1.get()
            self.comm.move_motors_to(unit, rotary_motor, right_z_axis_motor, left_z_axis_motor)
        else:
            pass
    
    def set_home_rotary_motor(self):
        if self.comm:
            self.comm.set_home_rotary_motor()
        else:
            pass
    
    def set_home_right_z_axis_motor(self):
        if self.comm:
            self.comm.set_home_right_z_axis_motor()
        else:
            pass
    
    def set_home_left_z_axis_motor(self):
        if self.comm:
            self.comm.set_home_left_z_axis_motor()
        else:
            pass

    def pid_experiment(self):
        if self.comm:
            pid_rotary_motor_direction = self.parent.control_frame.r1_c1.get()
            pid_rotary_motor_speed_rpm = self.parent.control_frame.r2_c1.get()
            if self.parent.control_frame.r3_c1.get() == "Yes":
                self.pid_target_load_lbf = self.parent.control_frame.r4_c1.get()
            else:   
                self.pid_target_load_lbf = 0
            pid_experiment_duration_hh_mm_ss = self.parent.control_frame.r5_c1.get()

            hh, mm, ss = map(int, pid_experiment_duration_hh_mm_ss.split(':'))
            self.pid_experiment_duration_ms = (hh*3600 + mm*60 + ss) * 1000

            self.comm.pid_experiment(pid_rotary_motor_direction, pid_rotary_motor_speed_rpm, self.pid_target_load_lbf, self.pid_experiment_duration_ms)
        else:
            pass

    def pid_experiment_terminate(self):
        if self.comm:
             self.comm.pid_experiment_terminate()
             self.experiment_data_count = 0
        else:
            pass